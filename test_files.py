import zipfile
import csv
from pypdf import PdfReader
from openpyxl import load_workbook
from settings_os import ARCHIVE
from io import TextIOWrapper


def test_csv_file():
    """Проверка содержимого csv-файла в архиве"""
    with zipfile.ZipFile(ARCHIVE, 'r', zipfile.ZIP_DEFLATED) as zip_archive:
        with zip_archive.open('file.csv', 'r', ) as csv_file:
            reader = list(csv.reader(TextIOWrapper(csv_file, 'utf-8-sig'), delimiter=';'))
            second_row = reader[1]

    assert second_row[1] == 'Dulce'  # проверка значения элемента во втором столбце второй строки
    assert second_row[2] == 'Abril'  # проверка значения элемента в третьем столбце второй строки


def test_pdf_file():
    """Проверка содержимого pdf-файла в архиве"""
    with zipfile.ZipFile(ARCHIVE, 'r', zipfile.ZIP_DEFLATED) as zip_archive:
        with zip_archive.open('file.pdf', 'r', ) as pdf_file:
            reader = PdfReader(pdf_file)
            sheets_count = len(reader.pages)

    assert sheets_count == 3


def test_xlsx_file():
    """Проверка содержимого xlsx-файла в архиве"""
    with zipfile.ZipFile(ARCHIVE, 'r', zipfile.ZIP_DEFLATED) as zip_archive:
        with zip_archive.open('file.xlsx', 'r', ) as xlsx_file:
            workbook = load_workbook(xlsx_file)
            worksheet = workbook.active
            anna_age = worksheet.cell(row=2, column=3).value
            person_name = worksheet.cell(row=5, column=5).value
            rows_count = worksheet.max_row

    assert anna_age == 18
    assert person_name == 'Tina'
    assert rows_count == 6

